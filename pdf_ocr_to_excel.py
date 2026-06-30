#!/usr/bin/env python3
"""
PDF OCR -> Excel 転記スクリプト
請求書PDFを読み取り、表の構成を保ったままExcelに転記する

使い方:
    python pdf_ocr_to_excel.py <PDFファイル> [出力Excelファイル]
    python pdf_ocr_to_excel.py invoice.pdf output.xlsx
"""

import io
import sys
import re
import threading
import tkinter as tk
from tkinter import filedialog, scrolledtext, ttk
from pathlib import Path

import numpy as np
import pdfplumber
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# --- OCRライブラリ可用性チェック ---
try:
    from paddleocr import PaddleOCR
    _PADDLE_AVAILABLE = True
except ImportError:
    _PADDLE_AVAILABLE = False

try:
    import pytesseract
    _TESSERACT_AVAILABLE = True
except ImportError:
    _TESSERACT_AVAILABLE = False

# pdf2image は PaddleOCR / Tesseract 両方で共用
try:
    from pdf2image import convert_from_path
    _PDF2IMAGE_AVAILABLE = True
except ImportError:
    _PDF2IMAGE_AVAILABLE = False


# ---------------------------------------------------------------------------
# 抽出レイヤー
# ---------------------------------------------------------------------------

def extract_with_pdfplumber(pdf_path: str) -> dict:
    """pdfplumber でテキストと表を抽出する（テキストPDF向け）"""
    data = {"raw_text": "", "_tables": [], "_words": []}
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            data["raw_text"] += (page.extract_text() or "") + "\n"
            data["_tables"].extend(page.extract_tables() or [])
            data["_words"].extend(page.extract_words() or [])
    return data


def _pdf_to_images(pdf_path: str):
    """PDFをPIL Imageのリストに変換する（DPI=300）"""
    if not _PDF2IMAGE_AVAILABLE:
        raise RuntimeError(
            "pdf2image が見つかりません: pip install pdf2image\n"
            "また Poppler のインストールも必要です。"
        )
    return convert_from_path(pdf_path, dpi=300)


def _group_into_lines(words: list, y_overlap_ratio: float = 0.4) -> list:
    """
    Y座標の重なりで word を行にグループ化し、X座標でソートして返す。
    y_overlap_ratio: 行高さに対して重なりがこの割合以上なら同一行とみなす
    """
    if not words:
        return []
    words = sorted(words, key=lambda w: w["top"])
    lines = [[words[0]]]

    for w in words[1:]:
        last = lines[-1]
        # 最終行の平均 top/bottom
        avg_top    = sum(x["top"]    for x in last) / len(last)
        avg_bottom = sum(x["bottom"] for x in last) / len(last)
        avg_height = max(avg_bottom - avg_top, 1)

        overlap = min(w["bottom"], avg_bottom) - max(w["top"], avg_top)
        if overlap >= avg_height * y_overlap_ratio:
            last.append(w)
        else:
            lines.append([w])

    return [sorted(line, key=lambda w: w["x0"]) for line in lines]


def _lines_to_table(lines_grouped: list) -> list:
    """
    行グループからテーブル構造（list[list[str]]）を構築する。
    列位置はX座標クラスタリングで推定する。
    """
    if not lines_grouped:
        return []

    # 全 word の x0 を収集 → 列境界をクラスタリング
    all_x0 = sorted(set(round(w["x0"] / 15) * 15 for line in lines_grouped for w in line))

    def col_index(x0):
        for i, cx in enumerate(all_x0):
            if abs(x0 - cx) <= 20:
                return i
        return len(all_x0) - 1

    n_cols = len(all_x0)
    table = []
    for line in lines_grouped:
        row = [""] * n_cols
        for w in line:
            ci = col_index(w["x0"])
            row[ci] = (row[ci] + " " + w["text"]).strip()
        # 末尾の空セルを取り除く
        while row and row[-1] == "":
            row.pop()
        if any(row):
            table.append(row)

    return table


def extract_with_paddleocr(pdf_path: str) -> dict:
    """PaddleOCR でOCR抽出する（スキャンPDF向け・メイン OCR エンジン）"""
    if not _PADDLE_AVAILABLE:
        raise RuntimeError(
            "PaddleOCR が見つかりません: pip install paddleocr paddlepaddle"
        )
    data = {"raw_text": "", "_tables": [], "_words": []}
    ocr = PaddleOCR(use_angle_cls=True, lang="japan", show_log=False)

    for img in _pdf_to_images(pdf_path):
        img_np = np.array(img)
        result = ocr.ocr(img_np, cls=True)
        if not result or not result[0]:
            continue

        words = []
        for line in result[0]:
            bbox, (text, conf) = line
            xs = [p[0] for p in bbox]
            ys = [p[1] for p in bbox]
            words.append({
                "text": text, "conf": conf,
                "x0": min(xs), "top": min(ys),
                "x1": max(xs), "bottom": max(ys),
            })

        lines_grouped = _group_into_lines(words)
        data["raw_text"] += "\n".join(
            " ".join(w["text"] for w in line) for line in lines_grouped
        ) + "\n"
        data["_words"].extend(words)

        table = _lines_to_table(lines_grouped)
        if table:
            data["_tables"].append(table)

    return data


def extract_with_tesseract(pdf_path: str) -> dict:
    """pdf2image + pytesseract でOCR抽出する（フォールバック）"""
    if not _TESSERACT_AVAILABLE:
        raise RuntimeError(
            "pytesseract が見つかりません: pip install pytesseract\n"
            "また Tesseract-OCR 本体と jpn.traineddata のインストールも必要です。"
        )
    data = {"raw_text": "", "_tables": [], "_words": []}
    for img in _pdf_to_images(pdf_path):
        data["raw_text"] += pytesseract.image_to_string(img, lang="jpn") + "\n"
    return data


# ---------------------------------------------------------------------------
# 解析レイヤー
# ---------------------------------------------------------------------------

def _num(s: str) -> str:
    """カンマ・記号を除いた数字文字列を返す"""
    return re.sub(r"[,，¥￥\s]", "", s or "")


def parse_invoice(data: dict) -> dict:
    """抽出データから請求書フィールドを解析する"""
    invoice = {
        "title": "請求書",
        "date": "",
        "recipient": "",
        "company_name": "",
        "address": "",
        "tel": "",
        "fax": "",
        "bank_name": "",
        "account_type": "",
        "account_number": "",
        "account_holder": "",
        "total_amount": "",
        "items": [],
        "subtotal": "",
        "tax": "",
        "grand_total": "",
        "notes": "",
    }

    text = data.get("raw_text", "")
    lines = [l.strip() for l in text.splitlines() if l.strip()]

    # --- テーブルから明細・集計を解析 ---
    item_section = False
    for table in data.get("_tables", []):
        for row in (table or []):
            if not row:
                continue
            cells = [c or "" for c in row]
            joined = "".join(cells)

            # ヘッダー行
            if re.search(r"品.?名|数量|単価|金額", joined):
                item_section = True
                continue

            # 集計行
            if re.search(r"小.?計", joined) and not re.search(r"品|洗|清", joined):
                nums = [_num(c) for c in cells if re.search(r"\d", c)]
                if nums:
                    invoice["subtotal"] = nums[-1]
                continue
            if "消費税" in joined:
                nums = [_num(c) for c in cells if re.search(r"\d", c)]
                if nums:
                    invoice["tax"] = nums[-1]
                continue
            if re.search(r"合.?計", joined) and not re.search(r"品|洗|清", joined):
                nums = [_num(c) for c in cells if re.search(r"\d", c)]
                if nums:
                    invoice["grand_total"] = nums[-1]
                continue
            if "備考" in joined:
                continue

            # 明細行
            if item_section and len(cells) >= 3:
                name = cells[0].strip()
                if not name or re.search(r"小計|消費税|合計|備考", name):
                    continue
                qty_raw  = cells[1].strip() if len(cells) > 1 else ""
                up_raw   = cells[2].strip() if len(cells) > 2 else ""
                amt_raw  = cells[3].strip() if len(cells) > 3 else ""
                note_raw = cells[4].strip() if len(cells) > 4 else ""

                qty = _num(qty_raw)
                up  = _num(up_raw)
                amt = _num(amt_raw)

                invoice["items"].append({
                    "name":       name,
                    "qty":        int(qty) if qty.isdigit() else qty_raw,
                    "unit_price": int(up)  if up.isdigit()  else up_raw,
                    "amount":     int(amt) if amt.isdigit() else amt_raw,
                    "note":       note_raw,
                })

    # --- テキストからヘッダー情報を解析 ---
    for i, line in enumerate(lines):
        # 日付
        m = re.search(r"(\d+)\s*年\s*(\d+)\s*月\s*(\d+)\s*日", line)
        if m and not invoice["date"]:
            invoice["date"] = f"{m.group(1)}年{m.group(2)}月{m.group(3)}日"

        # TEL / FAX
        m = re.search(r"TEL[.．\s]*([\d\-()（）]+)", line, re.IGNORECASE)
        if m:
            invoice["tel"] = m.group(1).strip()
        m = re.search(r"FAX[.．\s]*([\d\-()（）]+)", line, re.IGNORECASE)
        if m:
            invoice["fax"] = m.group(1).strip()

        # 銀行
        m = re.search(r"([^\s　]+銀行[^\s　]*)", line)
        if m and not invoice["bank_name"]:
            invoice["bank_name"] = m.group(1)

        # 口座情報
        m = re.search(r"預金種別[：:]\s*(.+)", line)
        if m:
            invoice["account_type"] = m.group(1).strip()
        m = re.search(r"口座番号[：:]\s*(\S+)", line)
        if m:
            invoice["account_number"] = m.group(1).strip()
        m = re.search(r"口座名義[：:]\s*(.+)", line)
        if m:
            invoice["account_holder"] = m.group(1).strip()

        # 御請求金額
        if re.search(r"御請求|請求金額", line):
            m = re.search(r"[¥￥]([\d,，]+)", line)
            if m:
                invoice["total_amount"] = _num(m.group(1))

        # 備考
        if re.match(r"備考[：:]?$", line):
            note_lines = lines[i+1:i+4]
            invoice["notes"] = " ".join(note_lines)

    # total_amount が未取得なら grand_total で代用
    if not invoice["total_amount"] and invoice["grand_total"]:
        invoice["total_amount"] = invoice["grand_total"]

    return invoice


# ---------------------------------------------------------------------------
# Excel書き出しレイヤー
# ---------------------------------------------------------------------------

_THIN   = Side(style="thin")
_MEDIUM = Side(style="medium")

def _border(t=None, b=None, l=None, r=None):
    return Border(top=t or _THIN, bottom=b or _THIN,
                  left=l or _THIN, right=r or _THIN)

_BORDER_ALL = _border()
_GRAY_FILL  = PatternFill(fill_type="solid", fgColor="DDDDDD")


def _cell(ws, row, col, value="", *, bold=False, size=11,
          align="left", valign="center", border=None, fill=None,
          number_format=None, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="MS Gothic", bold=bold, size=size)
    c.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
    if border is not None:
        c.border = border
    if fill is not None:
        c.fill = fill
    if number_format:
        c.number_format = number_format
    return c


def _merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


def _fill_border(ws, r1, c1, r2, c2):
    """マージ済みセル群に個別枠線を引く"""
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).border = _BORDER_ALL


def write_excel(invoice: dict, output_path: str):
    """請求書データを Excel に書き出す"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "請求書"

    # 列幅設定（列 A=1 〜 J=10）
    # A:余白, B:品名1, C:品名2, D:品名3, E:数量, F:単価, G:金額, H:摘要1, I:摘要2, J:余白
    col_widths = [2, 14, 7, 7, 7, 10, 12, 8, 8, 2]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # デフォルト行高さ
    ws.sheet_format.defaultRowHeight = 18

    # ----------------------------------------------------------------
    # タイトル（行1）
    # ----------------------------------------------------------------
    _merge(ws, 1, 2, 1, 9)
    _cell(ws, 1, 2, "請　　求　　書", bold=True, size=18, align="center")
    ws.row_dimensions[1].height = 30

    # ----------------------------------------------------------------
    # 日付（行2）
    # ----------------------------------------------------------------
    _merge(ws, 2, 6, 2, 9)
    _cell(ws, 2, 6, invoice.get("date") or "　　年　　月　　日", align="right")

    # ----------------------------------------------------------------
    # 宛先 / 会社情報（行3〜5）
    # ----------------------------------------------------------------
    ws.row_dimensions[3].height = 22
    _merge(ws, 3, 2, 3, 4)
    _cell(ws, 3, 2, invoice.get("recipient", ""), bold=True, size=13)
    _cell(ws, 3, 5, "様", bold=True, size=13)

    _merge(ws, 3, 6, 3, 9)
    _cell(ws, 3, 6, invoice.get("company_name", "(会社名)"), align="center")

    _merge(ws, 4, 6, 4, 9)
    _cell(ws, 4, 6, "〒　" + invoice.get("address", ""))

    _merge(ws, 5, 6, 5, 9)
    _cell(ws, 5, 6, invoice.get("address", "(住所)"))

    # ----------------------------------------------------------------
    # 下記のとおり / TEL / FAX（行6〜7）
    # ----------------------------------------------------------------
    _merge(ws, 6, 2, 6, 5)
    _cell(ws, 6, 2, "下記のとおりご請求いたします。")

    _merge(ws, 6, 6, 6, 9)
    _cell(ws, 6, 6, f"TEL. {invoice.get('tel', '')}")
    _merge(ws, 7, 6, 7, 9)
    _cell(ws, 7, 6, f"FAX. {invoice.get('fax', '')}")

    # ----------------------------------------------------------------
    # 振込先（行8〜12）
    # ----------------------------------------------------------------
    _merge(ws, 8, 2, 8, 5)
    _cell(ws, 8, 2, "（振込先）")
    _merge(ws, 9, 2, 9, 5)
    _cell(ws, 9, 2, invoice.get("bank_name", ""))
    _merge(ws, 10, 2, 10, 5)
    _cell(ws, 10, 2, f"預金種別：{invoice.get('account_type', '')}")
    _merge(ws, 11, 2, 11, 5)
    _cell(ws, 11, 2, f"口座番号：{invoice.get('account_number', '')}")
    _merge(ws, 12, 2, 12, 5)
    _cell(ws, 12, 2, f"口座名義：{invoice.get('account_holder', '')}")

    # 検印 / 担当者印（行9〜12, 列H・I）
    for r in range(9, 13):
        ws.row_dimensions[r].height = 22
    _cell(ws, 9, 8, "検　印",   align="center", border=_BORDER_ALL, fill=_GRAY_FILL)
    _cell(ws, 9, 9, "担当者印", align="center", border=_BORDER_ALL, fill=_GRAY_FILL)
    _merge(ws, 10, 8, 12, 8)
    _fill_border(ws, 10, 8, 12, 8)
    _merge(ws, 10, 9, 12, 9)
    _fill_border(ws, 10, 9, 12, 9)

    # ----------------------------------------------------------------
    # 御請求金額（行13）
    # ----------------------------------------------------------------
    ws.row_dimensions[13].height = 26
    _merge(ws, 13, 2, 13, 4)
    _cell(ws, 13, 2, "御請求金額", bold=True, size=13)

    total = invoice.get("total_amount") or invoice.get("grand_total", "")
    total_str = f"¥{int(total):,}-" if str(total).isdigit() else (f"¥{total}-" if total else "")
    _merge(ws, 13, 5, 13, 7)
    _cell(ws, 13, 5, total_str, bold=True, size=14, align="center")

    _merge(ws, 13, 8, 13, 9)
    _cell(ws, 13, 8, "（消費税込み）", size=9, align="center")

    # 区切り行（行14）
    ws.row_dimensions[14].height = 6

    # ----------------------------------------------------------------
    # 明細テーブルヘッダー（行15）
    # ----------------------------------------------------------------
    HDR = 15
    ws.row_dimensions[HDR].height = 22

    header_cols = [
        ("品　　名",    2, 4),
        ("数量",       5, 5),
        ("単価",       6, 6),
        ("金額",       7, 7),
        ("摘要",       8, 9),
    ]
    for label, c1, c2 in header_cols:
        if c1 != c2:
            _merge(ws, HDR, c1, HDR, c2)
        _cell(ws, HDR, c1, label, bold=True, align="center",
              border=_BORDER_ALL, fill=_GRAY_FILL)
        _fill_border(ws, HDR, c1, HDR, c2)

    # ----------------------------------------------------------------
    # 明細行（最大12行）
    # ----------------------------------------------------------------
    MAX_ITEM_ROWS = 12
    items = invoice.get("items", [])

    for i in range(MAX_ITEM_ROWS):
        r = HDR + 1 + i
        ws.row_dimensions[r].height = 18

        if i < len(items):
            it = items[i]
            _merge(ws, r, 2, r, 4)
            _cell(ws, r, 2, it.get("name", ""), border=_BORDER_ALL)
            _fill_border(ws, r, 2, r, 4)

            qty = it.get("qty", "")
            _cell(ws, r, 5, qty, align="center", border=_BORDER_ALL)

            up = it.get("unit_price", "")
            _cell(ws, r, 6, up, align="right", border=_BORDER_ALL,
                  number_format="#,##0")

            amt = it.get("amount", "")
            _cell(ws, r, 7, amt, align="right", border=_BORDER_ALL,
                  number_format="#,##0")

            note = it.get("note", "")
            _merge(ws, r, 8, r, 9)
            _cell(ws, r, 8, note, align="center", border=_BORDER_ALL)
            _fill_border(ws, r, 8, r, 9)
        else:
            # 空行
            _merge(ws, r, 2, r, 4)
            _fill_border(ws, r, 2, r, 4)
            _cell(ws, r, 5, "", border=_BORDER_ALL)
            _cell(ws, r, 6, "", border=_BORDER_ALL)
            _cell(ws, r, 7, "", border=_BORDER_ALL)
            _merge(ws, r, 8, r, 9)
            _fill_border(ws, r, 8, r, 9)

    # ----------------------------------------------------------------
    # 集計行（小計 / 消費税等 / 合計）
    # ----------------------------------------------------------------
    SUM_START = HDR + 1 + MAX_ITEM_ROWS

    summaries = [
        ("小　　計", invoice.get("subtotal", ""),    False),
        ("消費税等", invoice.get("tax", ""),         False),
        ("合　　計", invoice.get("grand_total", ""), True),
    ]
    for idx, (label, value, is_bold) in enumerate(summaries):
        r = SUM_START + idx
        ws.row_dimensions[r].height = 20

        _merge(ws, r, 2, r, 6)
        _cell(ws, r, 2, label, bold=is_bold, align="center", border=_BORDER_ALL)
        _fill_border(ws, r, 2, r, 6)

        v = str(value).replace(",", "")
        cell_val = int(v) if v.isdigit() else value
        _cell(ws, r, 7, cell_val, bold=is_bold, align="right",
              border=_BORDER_ALL, number_format="#,##0")

        _merge(ws, r, 8, r, 9)
        _fill_border(ws, r, 8, r, 9)

    # ----------------------------------------------------------------
    # 備考欄
    # ----------------------------------------------------------------
    NR = SUM_START + len(summaries)
    ws.row_dimensions[NR].height = 18
    ws.row_dimensions[NR + 1].height = 45
    ws.row_dimensions[NR + 2].height = 45

    _merge(ws, NR, 2, NR, 9)
    _cell(ws, NR, 2, "備考：", border=_BORDER_ALL)
    _fill_border(ws, NR, 2, NR, 9)

    _merge(ws, NR + 1, 2, NR + 2, 9)
    _cell(ws, NR + 1, 2, invoice.get("notes", ""), border=_BORDER_ALL, wrap=True)
    _fill_border(ws, NR + 1, 2, NR + 2, 9)

    wb.save(output_path)
    print(f"[完了] Excel出力: {output_path}")


# ---------------------------------------------------------------------------
# 変換処理（GUI / CLI 共通）
# ---------------------------------------------------------------------------

def run_conversion(pdf_path: str, output_path: str, log=print):
    """PDF → Excel 変換の本体。log に呼び出し可能オブジェクトを渡すとそこへ出力する。"""
    if not Path(pdf_path).exists():
        raise FileNotFoundError(f"ファイルが見つかりません: {pdf_path}")

    log(f"[1/3] PDF読み取り中: {pdf_path}")
    data = extract_with_pdfplumber(pdf_path)

    if len(data.get("raw_text", "").strip()) < 30:
        if _PADDLE_AVAILABLE:
            log("      → PaddleOCR で再抽出します...")
            data = extract_with_paddleocr(pdf_path)
        elif _TESSERACT_AVAILABLE:
            log("      → PaddleOCR 未インストール。Tesseract にフォールバックします...")
            data = extract_with_tesseract(pdf_path)
        else:
            log("[警告] OCRエンジンが見つかりません。テキスト抽出結果のみで続行します。")

    log("[2/3] データ解析中...")
    invoice = parse_invoice(data)

    log(f"      日付      : {invoice['date'] or '（未検出）'}")
    log(f"      会社名    : {invoice['company_name'] or '（未検出）'}")
    log(f"      請求金額  : {invoice['total_amount'] or '（未検出）'}")
    log(f"      明細件数  : {len(invoice['items'])} 件")
    for it in invoice["items"]:
        log(f"        - {it['name']}  数量:{it['qty']}  単価:{it['unit_price']}  金額:{it['amount']}")

    log("[3/3] Excel書き出し中...")
    write_excel(invoice, output_path)


# ---------------------------------------------------------------------------
# GUI
# ---------------------------------------------------------------------------

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("PDF OCR → Excel 転記")
        self.resizable(False, False)
        self._build_ui()

    # ---- UI構築 ----

    def _build_ui(self):
        PAD = dict(padx=10, pady=4)

        # --- 入力PDF ---
        frm_in = ttk.LabelFrame(self, text="入力 PDF ファイル")
        frm_in.grid(row=0, column=0, sticky="ew", padx=12, pady=(12, 4))

        self._pdf_var = tk.StringVar()
        ttk.Entry(frm_in, textvariable=self._pdf_var, width=54).grid(
            row=0, column=0, **PAD)
        ttk.Button(frm_in, text="参照…", command=self._browse_pdf).grid(
            row=0, column=1, padx=(0, 8))

        # --- 出力Excel ---
        frm_out = ttk.LabelFrame(self, text="出力 Excel ファイル")
        frm_out.grid(row=1, column=0, sticky="ew", padx=12, pady=4)

        self._xlsx_var = tk.StringVar()
        ttk.Entry(frm_out, textvariable=self._xlsx_var, width=54).grid(
            row=0, column=0, **PAD)
        ttk.Button(frm_out, text="参照…", command=self._browse_xlsx).grid(
            row=0, column=1, padx=(0, 8))

        # --- 実行ボタン ---
        self._run_btn = ttk.Button(self, text="変換実行", command=self._start,
                                   style="Accent.TButton")
        self._run_btn.grid(row=2, column=0, pady=8)

        # --- ログエリア ---
        frm_log = ttk.LabelFrame(self, text="ログ")
        frm_log.grid(row=3, column=0, sticky="nsew", padx=12, pady=(0, 4))

        self._log_box = scrolledtext.ScrolledText(
            frm_log, width=72, height=14, state="disabled",
            font=("Consolas", 9), wrap="word")
        self._log_box.pack(padx=6, pady=6)

        # --- ステータスバー ---
        self._status_var = tk.StringVar(value="PDFファイルを選択してください")
        ttk.Label(self, textvariable=self._status_var, anchor="w",
                  relief="sunken").grid(row=4, column=0, sticky="ew",
                                        padx=12, pady=(0, 10))

        # --- プログレスバー ---
        self._progress = ttk.Progressbar(self, mode="indeterminate")
        self._progress.grid(row=5, column=0, sticky="ew", padx=12, pady=(0, 10))

    # ---- ファイル選択 ----

    def _browse_pdf(self):
        path = filedialog.askopenfilename(
            title="PDFファイルを選択",
            filetypes=[("PDF ファイル", "*.pdf"), ("すべてのファイル", "*.*")])
        if not path:
            return
        self._pdf_var.set(path)
        # 出力パスを自動補完（未入力のとき）
        if not self._xlsx_var.get():
            default_out = str(Path(path).with_suffix("")) + "_output.xlsx"
            self._xlsx_var.set(default_out)
        self._status_var.set("入力ファイル設定済み")

    def _browse_xlsx(self):
        init_dir  = str(Path(self._pdf_var.get()).parent) if self._pdf_var.get() else "."
        init_file = Path(self._xlsx_var.get()).name if self._xlsx_var.get() else "output.xlsx"
        path = filedialog.asksaveasfilename(
            title="保存先を選択",
            initialdir=init_dir,
            initialfile=init_file,
            defaultextension=".xlsx",
            filetypes=[("Excel ファイル", "*.xlsx"), ("すべてのファイル", "*.*")])
        if path:
            self._xlsx_var.set(path)

    # ---- 変換実行 ----

    def _start(self):
        pdf  = self._pdf_var.get().strip()
        xlsx = self._xlsx_var.get().strip()

        if not pdf:
            self._status_var.set("[エラー] PDFファイルを選択してください")
            return
        if not xlsx:
            self._status_var.set("[エラー] 出力ファイルを指定してください")
            return

        self._run_btn.config(state="disabled")
        self._progress.start(12)
        self._status_var.set("変換中…")
        self._log_clear()

        threading.Thread(target=self._worker, args=(pdf, xlsx), daemon=True).start()

    def _worker(self, pdf: str, xlsx: str):
        try:
            run_conversion(pdf, xlsx, log=self._log)
            self.after(0, self._on_success, xlsx)
        except Exception as exc:
            self.after(0, self._on_error, str(exc))

    # ---- コールバック ----

    def _on_success(self, xlsx: str):
        self._progress.stop()
        self._run_btn.config(state="normal")
        self._status_var.set(f"[完了] {xlsx}")
        self._log(f"\n[完了] 出力: {xlsx}")

    def _on_error(self, msg: str):
        self._progress.stop()
        self._run_btn.config(state="normal")
        self._status_var.set(f"[エラー] {msg}")
        self._log(f"\n[エラー] {msg}")

    # ---- ログユーティリティ ----

    def _log(self, text: str):
        """スレッドセーフなログ追記"""
        def _append():
            self._log_box.config(state="normal")
            self._log_box.insert("end", text + "\n")
            self._log_box.see("end")
            self._log_box.config(state="disabled")
        self.after(0, _append)

    def _log_clear(self):
        self._log_box.config(state="normal")
        self._log_box.delete("1.0", "end")
        self._log_box.config(state="disabled")


# ---------------------------------------------------------------------------
# エントリーポイント
# ---------------------------------------------------------------------------

def main():
    # 引数あり → CLIモード
    if len(sys.argv) >= 2:
        pdf_path    = sys.argv[1]
        output_path = sys.argv[2] if len(sys.argv) > 2 else Path(sys.argv[1]).stem + "_output.xlsx"
        try:
            run_conversion(pdf_path, output_path)
        except Exception as exc:
            print(f"[エラー] {exc}", file=sys.stderr)
            sys.exit(1)
        return

    # 引数なし → GUIモード
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()
